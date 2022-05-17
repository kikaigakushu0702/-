#%%
import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib 

wb = openpyxl.load_workbook('shain_data.xlsx')
ws = wb.active
titles = ws[1]
ws.delete_rows(1)


values_name = []
values_age = []
values_address = []
values_group = []
values_zan = []
values_visualtime =[]


for cell in ws['A']:
    values_name.append(cell.value)
    
for cell in ws['B']:
    values_age.append(cell.value)

for cell in ws['C']:
    values_address.append(cell.value)

for cell in ws['D']:
    values_group.append(cell.value)

for cell in ws['E']:
    values_zan.append(cell.value)

for cell in ws['F']:
    values_visualtime.append(cell.value)



fig,ax = plt.subplots(2,2, facecolor='white',figsize=(20,14),)


a = ax[0,0]
b = ax[0,1]
c = ax[1,0]
d = ax[1,1]


a.bar(values_name,values_zan)
a.set_title("bar graph",)
a.set_xlabel((titles[0].value))
a.set_ylabel((titles[4].value),rotation='horizontal',labelpad=35)
a.set_yticks([0,5,10,15,20,25,30,35,40,45,])

b.hist(values_visualtime, bins=5)
b.set_title("hist")
b.set_xlabel((titles[5].value))
b.set_ylabel(("区間"),rotation='horizontal',labelpad=35)

c.scatter(values_zan,values_group)
c.set_xlabel((titles[4].value))
c.set_ylabel((titles[3].value),rotation='horizontal',labelpad=35)
c.set_title("scatter1")

d.scatter(values_zan,values_visualtime)
d.set_xlabel((titles[4].value))
d.set_ylabel((titles[5].value),rotation='horizontal',labelpad=35)
d.set_title("scatter2")


fig.savefig("img.png")


    













# %%
